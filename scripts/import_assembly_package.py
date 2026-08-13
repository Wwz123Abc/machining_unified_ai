"""Import a STEP assembly package with BOM and drawing references."""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from machining_unified.config.paths import (  # noqa: E402
    ASSEMBLY_PACKAGES_DIR,
    CAD_SAMPLES_DIR,
)


# 原始资料包会复制到 data/enterprise/assembly_packages 以便审计；
# 仅其装配 STEP 会放入 data/enterprise/cad_samples 参与常规几何索引。


def normalized_part_id(value: str) -> str:
    # 工程图与 BOM 系统可能增加三位产品前缀或版本后缀。
    # 标准化仅生成谨慎的关联键，不覆盖原始编码。
    text = re.sub(r"\.[^.]+$", "", value.upper())
    text = re.sub(r"^[0-9]{3}(?=DTXT)", "", text)
    return re.sub(r"([0-9]{3})[A-Z]$", r"\1", text)


def read_bom(path: Path) -> list[dict[str, Any]]:
    # 查找真实表头行而不假设固定表格布局，可处理 BOM 表上方的标题块，
    # 同时保留原始字段值。
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next(index for index, row in enumerate(rows) if any(value and "P/N" in str(value) for value in row))
    headers = [str(value or "").replace("\n", " ").strip() for value in rows[header_index]]
    items = []
    for row in rows[header_index + 1 :]:
        values = dict(zip(headers, row))
        part_id = values.get("P/N 物料编码") or values.get("P/N 物料编码 ")
        if not part_id:
            continue
        items.append(
            {
                "part_id": str(part_id).strip(),
                "normalized_part_id": normalized_part_id(str(part_id)),
                "revision": str(values.get("Rev. 版本") or "").strip(),
                "name": str(values.get("Description 名称") or "").strip(),
                "specification": str(values.get("Specification 物料描述") or "").strip(),
                "drawing_no": str(values.get("Model No. 型号/图号") or "").strip(),
                "material": str(values.get("材料") or "").strip(),
                "surface_treatment": str(values.get("表处") or "").strip(),
                "quantity": values.get("QTY 数量"),
                "unit": str(values.get("Unit 单位") or "").strip(),
            }
        )
    return items


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source_dir", type=Path)
    args = parser.parse_args()
    source_dir = args.source_dir.resolve()
    step_files = [path for path in source_dir.rglob("*") if path.is_file() and path.suffix.lower() in {".step", ".stp"}]
    if len(step_files) != 1:
        raise SystemExit(f"Expected exactly one assembly STEP, found {len(step_files)}")
    assembly_step = step_files[0]
    assembly_id = assembly_step.stem
    bom_files = list((source_dir / "零件明细表").glob("*.xlsx"))
    if len(bom_files) != 1:
        raise SystemExit("Expected one detailed BOM workbook under 零件明细表")
    bom_items = read_bom(bom_files[0])
    package_dir = ASSEMBLY_PACKAGES_DIR / assembly_id
    package_dir.mkdir(parents=True, exist_ok=True)
    # 先复制原始资料文件，确保清单链接始终能在本项目内解析。
    for source in source_dir.rglob("*"):
        if source.is_file():
            target = package_dir / source.relative_to(source_dir)
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
    catalog_step_dir = CAD_SAMPLES_DIR / "assemblies" / assembly_id
    catalog_step_dir.mkdir(parents=True, exist_ok=True)
    catalog_step = catalog_step_dir / assembly_step.name
    shutil.copy2(assembly_step, catalog_step)
    drawings = {normalized_part_id(path.stem): str(path.relative_to(ROOT)).replace("\\", "/") for path in (package_dir / "工程图").glob("*.pdf")}
    for path in (package_dir / "工程图").glob("*.PDF"):
        drawings[normalized_part_id(path.stem)] = str(path.relative_to(ROOT)).replace("\\", "/")
    # 通过标准化编码关联工程图。找不到图纸时保持 None，
    # 不会根据相似文件名猜测关联关系。
    for item in bom_items:
        item["drawing_file"] = drawings.get(item["normalized_part_id"])
    manifest = {
        "assembly_id": assembly_id,
        "normalized_assembly_id": normalized_part_id(assembly_id),
        "assembly_step": str(catalog_step.relative_to(ROOT)).replace("\\", "/"),
        "source_package": str(package_dir.relative_to(ROOT)).replace("\\", "/"),
        "bom_file": str((package_dir / bom_files[0].relative_to(source_dir)).relative_to(ROOT)).replace("\\", "/"),
        "component_count": len(bom_items),
        "bom_items": bom_items,
    }
    (package_dir / "assembly_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Imported assembly {assembly_id}: {len(bom_items)} BOM items, {sum(1 for item in bom_items if item['drawing_file'])} linked drawings")


if __name__ == "__main__":
    main()
